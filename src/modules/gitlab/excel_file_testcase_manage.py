import shutil
import os
from openpyxl import load_workbook
from datetime import datetime

def create_testcase_file(iss_number, project, folder_name, file_name):
    print("create_testcase_file", iss_number, project, folder_name, file_name)
    path_dst_file_tmp = "D:\\Testcase\\RPA\\{0}\\{1}\\{2}.xlsx"
    path_dst_folder_tmp = "D:\\Testcase\\RPA\\{0}\\{1}"
    path_src = ".\\TEMPLATE\\Testcase-template-{0}.xlsx"
    src = str(path_src.format(project)), 
    path_file_dst = str(path_dst_file_tmp.format(project, folder_name, "".join(file_name)))
    path_folder_dst = str(path_dst_folder_tmp.format(project, folder_name))
    print(src[0], path_file_dst)

    if os.path.exists(path_folder_dst):
        print("Folder is exist: ", path_folder_dst)
    else:
        print("Folder is not exist!", path_folder_dst)
        os.makedirs(path_folder_dst)
    shutil.copy(src[0], path_file_dst)
    
    return path_file_dst

def update_file_testcase(path, iss_test_number, issue_desc, test_scenario):
    print("Update file testcase", path)
    
    #load excel file
    workbook = load_workbook(filename=path)
    
    #open workbook
    sheet = workbook.active
    
    #modify the desired cell
    sheet["C1"] = iss_test_number
    sheet["F1"] = issue_desc
    sheet["B14"] = test_scenario
    
    #save the file
    workbook.save(path)

def update_finish_date_file_testcase(path):
    print("Update finish date file testcase", path)
    
    #load excel file
    workbook = load_workbook(filename=path)
    
    #open workbook
    sheet = workbook.active

    # Merge cell F6 và F7
    sheet.merge_cells('F6:G6')

    # Gán giá trị ngày hiện tại theo format DD/MM/YYYY
    today_str = datetime.now().strftime("%d/%m/%Y")
    sheet["F6"] = today_str

    # save the file
    workbook.save(path)